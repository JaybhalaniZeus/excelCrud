from openpyxl import load_workbook

def update_excel_cell(file_path, sheet_name, row, column, new_value):
    try:
        # Load the workbook
        wb = load_workbook(filename=file_path)
        
        # Select the active worksheet
        ws = wb[sheet_name]
        
        # Update the value in the specified cell
        ws.cell(row=row, column=column, value=new_value)
        
        # Save the workbook
        wb.save(file_path)
        
        print(f"Data in cell ({row}, {column}) updated successfully.")
    
    except Exception as e:
        print("Error updating data in Excel file:", e)

def read_excel_cell(file_path, sheet_name, row, column):
    try:
        # Load the workbook
        wb = load_workbook(filename=file_path)
        
        # Select the active worksheet
        ws = wb[sheet_name]
        
        # Read the value from the specified cell
        cell_value = ws.cell(row=row, column=column).value
        
        print(f"Value from specified cell ({row}, {column}): {cell_value}")
    
    except Exception as e:
        print("Error reading data from Excel file:", e)

# Specify the path to the Excel file
excel_file_path = r"D:\excel\task.xlsx"

# Specify the name of the worksheet
sheet_name = "Sheet1"  # Example sheet name

# Specify the row and column to update the cell data
update_row_number = 3  # Example row number for update
update_column_number = 2  # Example column number for update

# Specify the new value to update the cell with
new_cell_value = "JAHAN"  # Example new value for update

# Call the function to update data in the specified cell
update_excel_cell(excel_file_path, sheet_name, update_row_number, update_column_number, new_cell_value)

# Specify the row and column to read the cell data
read_row_number = 4  # Example row number for read
read_column_number = 3  # Example column number for read

# Call the function to read data from the specified cell
read_excel_cell(excel_file_path, sheet_name, read_row_number, read_column_number)
